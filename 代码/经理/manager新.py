import datetime
import os
import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
import sys
from COMMONMenu import Ui_HotelMenu

def date_sort3(ls):
    # 日期排序
    # 用了冒泡排序来排序，其他方法效果一样
    for j in range(len(ls) - 1):
        for i in range(len(ls) - j - 1):
            try:
                lower = datetime.datetime.strptime(ls[i], '%Y-%m-%d %H:%M:%S')
                upper = datetime.datetime.strptime(ls[i + 1], '%Y-%m-%d %H:%M:%S')
                if lower > upper:
                    ls[i], ls[i + 1] = ls[i + 1], ls[i]
            except:
                print(ls[i])
    return ls


def get_week_num(dayTime): # '2020/06/14 12:00:00'
    # 返回指定时间属于第几周
    # datetime.date(2017, 12, 31).isocalendar()  (2017, 52, 7)  2017 年第 52周 周日
    dayTime = dayTime.split()[0].split('/')
    y = int(dayTime[0])
    m = int(dayTime[1])
    d = int(dayTime[2])
    return datetime.date(y, m, d).isocalendar()


# 获取某行所有值
def getRowValues(sheet, row):
    columns = sheet.max_column
    rowdata = []
    for i in range(1, columns + 1):
        cellvalue = sheet.cell(row=row, column=i).value
        rowdata.append(cellvalue)

    rowdata[-1] = str(rowdata[-1]).replace('-','/')
    return rowdata


# 写入数据
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)


def process_year(data_path):
    wb = openpyxl.load_workbook(data_path)  # 加载工作薄
    sheet = wb['Sheet1']
    # print(sheet)

    time_dict = {}  # 用于每一行的时间的原始排序  {'2019-04-08 10:22:05': 1, '2020-05-08 04:34:46': 2, '2020-05-22 16:31:16': 3, '2020-11-12 14:34:16': 4, '2021-03-05 23:39:16': 5}
    time_list = []  # 保存现有的时间点用于排序
    time_now = sheet["I"][1:]  # 第一行是名称，不取来计算
    for index, data in enumerate(time_now):
        sheet_I_time = data.value
        if sheet_I_time: # 去除最后面的空行
            sheet_I_time = str(sheet_I_time)
            time_list.append(sheet_I_time)
            time_dict[sheet_I_time] = index + 2  # 从2行开始，第一行是名称

    sort_time_list = date_sort3(time_list)  # 如果总的数据时间是从古到今排序的，可以不用排序

    # 获取年表
    current_year = sort_time_list[0]
    current_year = int(current_year.split('-')[0])  # int: 2019
    current_year_start_index = 0

    len_sort_time_list = len(sort_time_list)
    for index in range(len_sort_time_list):

        next_year = int(sort_time_list[min(index+1, len_sort_time_list-1)].split('-')[0])
        if next_year == current_year and index != len_sort_time_list - 1:  # 如果还是同一年，继续查找
            continue
        else:  # 不为同一年的时候
            book_name_xlsx = './years/%s.xlsx' % current_year
            sheet_name_xlsx = 'Sheet1'

            all_year_data = []
            all_year_data.append(getRowValues(sheet, 1))  # 获取第一行的数据
            for i in sort_time_list[current_year_start_index: index + 1]:
                temp_row_data = getRowValues(sheet, time_dict[i])
                all_year_data.append(temp_row_data)

            write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, all_year_data)

            # 更新年份开始下标和当前年份
            current_year = next_year
            current_year_start_index = index + 1


def process_month(data_path, year_num):
    wb = openpyxl.load_workbook(data_path)  # 加载工作薄
    sheet = wb['Sheet1']

    time_dict = {}  # 用于每一行的时间的原始排序  {'2019-04-08 10:22:05': 1, '2020-05-08 04:34:46': 2, '2020-05-22 16:31:16': 3, '2020-11-12 14:34:16': 4, '2021-03-05 23:39:16': 5}
    time_list = []  # 保存现有的时间点用于排序
    time_now = sheet["I"][1:]  # 第一行是名称，不取来计算
    for index, data in enumerate(time_now):
        sheet_I_time = str(data.value)
        time_list.append(sheet_I_time)
        time_dict[sheet_I_time] = index + 2  # 从2行开始，第一行是名称

    # sort_time_list = date_sort3(time_list)
    sort_time_list = time_list  # 年表已经排序了

    # 获取月表
    current_month = sort_time_list[0]
    current_month = int(current_month.split('/')[1])
    current_month_start_index = 0

    len_sort_time_list = len(sort_time_list)
    for index in range(len_sort_time_list):

        next_month = int(sort_time_list[min(index+1, len_sort_time_list-1)].split('/')[1])
        if next_month == current_month and index != len_sort_time_list - 1:  # 如果还是同一月，继续查找
            continue
        else:  # 不为同一月份的时候
            book_name_xlsx = './months/%s-%s.xlsx' % (year_num, current_month)
            sheet_name_xlsx = 'Sheet1'

            all_year_data = []
            all_year_data.append(getRowValues(sheet, 1))  # 获取第一行的数据
            for i in sort_time_list[current_month_start_index: index + 1]:
                temp_row_data = getRowValues(sheet, time_dict[i])
                all_year_data.append(temp_row_data)

            write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, all_year_data)

            # 更新年份开始下标和当前月份
            current_month = next_month
            current_month_start_index = index + 1


def process_week(data_path, year_num):
    wb = openpyxl.load_workbook(data_path)  # 加载工作薄
    sheet = wb['Sheet1']
    # print(sheet)

    time_dict = {}  # 用于每一行的时间的原始排序  {'2019-04-08 10:22:05': 1, '2020-05-08 04:34:46': 2, '2020-05-22 16:31:16': 3, '2020-11-12 14:34:16': 4, '2021-03-05 23:39:16': 5}
    time_list = []  # 保存现有的时间点用于排序
    time_now = sheet["I"][1:]  # 第一行是名称，不取来计算
    for index, data in enumerate(time_now):
        sheet_I_time = str(data.value)
        time_list.append(sheet_I_time)
        time_dict[sheet_I_time] = index + 2  # 从2行开始，第一行是名称

    sort_time_list = time_list  # 年表已经排序了,故不用排序

    current_week = get_week_num(sort_time_list[0])[1]
    current_week_start_index = 0

    len_sort_time_list = len(sort_time_list)
    for index in range(len_sort_time_list):

        next_week = get_week_num(sort_time_list[min(index + 1, len_sort_time_list - 1)])[1]
        if next_week == current_week and index != len_sort_time_list - 1:
            continue
        else:
            book_name_xlsx = './weeks/%s-%s_week.xlsx' % (year_num, current_week)
            sheet_name_xlsx = 'Sheet1'

            all_year_data = []
            all_year_data.append(getRowValues(sheet, 1))  # 获取第一行的数据
            for i in sort_time_list[current_week_start_index: index + 1]:
                temp_row_data = getRowValues(sheet, time_dict[i])
                all_year_data.append(temp_row_data)

            write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, all_year_data)

            # 更新年份开始下标和当前周数
            current_week = next_week
            current_week_start_index = index + 1


def process_day(data_path, year_num, month):
    wb = openpyxl.load_workbook(data_path)  # 加载工作薄
    sheet = wb['Sheet1']

    time_dict = {}  # 用于每一行的时间的原始排序  {'2019-04-08 10:22:05': 1, '2020-05-08 04:34:46': 2, '2020-05-22 16:31:16': 3, '2020-11-12 14:34:16': 4, '2021-03-05 23:39:16': 5}
    time_list = []  # 保存现有的时间点用于排序
    time_now = sheet["I"][1:]  # 第一行是名称，不取来计算
    for index, data in enumerate(time_now):
        sheet_I_time = str(data.value)
        time_list.append(sheet_I_time)
        time_dict[sheet_I_time] = index + 2

    sort_time_list = time_list
    current_day = sort_time_list[0]
    current_day = int(current_day.split()[0].split('/')[-1])
    current_day_start_index = 0

    len_sort_time_list = len(sort_time_list)
    for index in range(len_sort_time_list):

        next_month = int(sort_time_list[min(index+1, len_sort_time_list-1)].split()[0].split('/')[-1])
        if next_month == current_day and index != len_sort_time_list - 1:
            continue
        else:
            book_name_xlsx = './days/%s-%s-%s.xlsx' % (year_num, month, current_day)
            sheet_name_xlsx = 'Sheet1'

            all_year_data = []
            all_year_data.append(getRowValues(sheet, 1))  # 获取第一行的数据
            for i in sort_time_list[current_day_start_index: index + 1]:
                temp_row_data = getRowValues(sheet, time_dict[i])
                all_year_data.append(temp_row_data)

            write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, all_year_data)

            current_day = next_month
            current_day_start_index = index + 1


def main(data_path):
    if not os.path.exists('./years/'):
        os.mkdir('./years/')

    if not os.path.exists('./months/'):
        os.mkdir('./months/')

    if not os.path.exists('./weeks/'):
        os.mkdir('./weeks/')

    if not os.path.exists('./days/'):
        os.mkdir('./days/')

    process_year(data_path)

    for i in os.listdir('./years/'):
        year_num = int(i.split('.')[0])
        process_month('./years/' + i, year_num)
        process_week('./years/' + i, year_num)

    for i in os.listdir('./months/'):
        temp = i.split('.')[0].split('-')
        year_num = int(temp[0])
        month_num = int(temp[1])
        process_day('./months/'+i, year_num, month_num)

    ui.textBrowser_admin.append("Done! All the reports can be found in the directory!")

def PrintReportButton(self):
    main('./输入1.xlsx')

if __name__ == '__main__':
    # create the app
    app = QtWidgets.QApplication(sys.argv)

    # create the form and init UI
    HotelMenu = QtWidgets.QMainWindow()
    ui = Ui_HotelMenu()
    ui.setupUi(HotelMenu)
    HotelMenu.show()

    # create hook logic
    ui.ButtonPrint.clicked.connect(PrintReportButton)


    # run main loop
    sys.exit(app.exec_())
    # main('./输入1.xlsx')  # 输入总的数据，每年月周日在各自的文件夹中，如果表的时间的顺序是对的可以注释掉排序
    # process_week('./years/2020.xlsx', 2020)
    # process_day('./months/2020-5.xlsx', 2020, 5)
    # process_year('./输入2.xlsx')
    # process_month('./years/2020.xlsx', 2020)
    # process_week('./years/2020.xlsx', 2020)
    # print(get_week_num('2020/6/10 10:00:00'))