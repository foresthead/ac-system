from PyQt5 import QtCore, QtGui, QtWidgets
import sys
from COMMONMenu import Ui_HotelMenu
import os
import openpyxl

# RDR is Request Detail Records == 详单
# Invoice == 账单


# controller will tell classes what to do AND NOW it's also printing values

class ReceptionController:
    def __init__(self, RDRorI, OutputType, RoomID):
        self.RDRorI = RDRorI  # this attribute will tell the class, what did user choose - invoice or DR
        self.OutputType = OutputType
        self.ID = RoomID
        self.ThisRDR = ServiceRDR(self.ID)
        self.ThisInvoice = ServiceInvoice(self.ID)


    # this method is the constructor of RDR and invoice.
    def commandCreateList(self):
        if self.RDRorI == '1':

            # first, we need to add Number of RDR (+1)
            # self.book = openpyxl.load_workbook('./输入1.xlsx')  # database loading
            # self.sheet = self.book.active
            # self.RoomRow = (int(self.RoomID) + 1)

            # a = (self.sheet.cell(row=self.RoomRow, column=6).value)
            # print(a)


            RDRContent = self.ThisRDR.createList()
            self.Print(RDRContent)
        elif self.RDRorI == '2':
            InvoiceContent = self.ThisInvoice.createInvoiceData()
            self.Print(InvoiceContent)
        else:
            print("Processing error")


    # this method contains sends a printing command
    def Print(self, ListContent):
        self.ListContent = ListContent
        if self.RDRorI == '1':
            if self.OutputType == '1':
                self.PutOnScreen(ListContent)
            elif self.OutputType == '2':
                self.PutInTxt(ListContent)
            else:
                print("Printer error")

        elif self.RDRorI == '2':
            if self.OutputType == '1':
                self.PutOnScreen(ListContent)
            else:
                print("Printer error")
        else:
            print('print error')

    # These methods USED TO BE a part of a printer class. But it's not working with GUI I don't know why

    def PutOnScreen(self, ListContent):
        # ui.textBrowser.setText(ListContent)
        for i in range(len(ListContent)):  # output every string of our list_RDR
            # print(ListContent[i])
            ui.textBrowser_reception.append(ListContent[i])
        ui.textBrowser_reception.append("=================================")

    def PutInTxt(self, ListContent):
        import webbrowser
        with open('detailed_record.txt', 'w') as printfile:
            for line in ListContent:
                printfile.write(line + '\n')
            webbrowser.open("detailed_record.txt")



class Database:
    def __init__(self, RoomID):
        import datetime
        self.RoomID = RoomID
        self.book1 = openpyxl.load_workbook('./ACC_Database.xlsx')  # book1 is the ACC database, book2 is shuru1
        self.sheet1 = self.book1.active

        self.book2 = openpyxl.load_workbook('./输入1.xlsx')
        self.sheet2 = self.book2.active

        self.RoomRow = (int(self.RoomID) + 1)               # +1 cuz the first row is the column titles

        #getting values from book1
        self.FanSpeed = self.sheet1.cell(row=self.RoomRow, column=4).value
        # self.FeeRate = self.sheet1.cell(row=self.RoomRow, column=5).value
        self.Temperature = self.sheet1.cell(row=self.RoomRow, column=3).value
        self.CheckInTime = self.sheet1.cell(row=self.RoomRow, column=2).value
        self.temp_lowLimit = self.sheet1.cell(row=2, column=7).value
        if self.FanSpeed == 1:
            self.NewFeeRate = self.sheet1.cell(row=2, column=10).value
        elif self.FanSpeed == 2:
            self.NewFeeRate = self.sheet1.cell(row=2, column=9).value
        else:
            self.NewFeeRate = self.sheet1.cell(row=2, column=8).value
        self.Fee =  self.NewFeeRate * (self.Temperature - self.temp_lowLimit)
        self.TimeNow = datetime.datetime.now()

        # putting values in book2
        self.sheet1.cell(row=self.RoomRow, column=5).value = self.NewFeeRate

        # putting values in book2
        self.sheet2.cell(row=self.RoomRow, column=9).value = self.TimeNow

        self.NewTotalFee = (self.sheet2.cell(row=self.RoomRow, column=4).value) + int(self.Fee)
        self.sheet2.cell(row=self.RoomRow, column=4).value = self.NewTotalFee


        self.book1.save('./ACC_Database.xlsx')
        self.book2.save('./输入1.xlsx')

##################################################################################
#                                                                                #
#                             RDR RELATED PART                                   #
#                                                                                #
##################################################################################


class ServiceRDR:   # This class made for creating DR that we can output on screen or print to file.
    def __init__(self, ID):
        self.RoomID: int = ID
        self.ValuesSource = Database(self.RoomID)
        self.book = openpyxl.load_workbook('./输入1.xlsx')  # database loading
        self.sheet = self.book.active
        self.RoomRow = (int(self.RoomID) + 1)

    # taking values using getValues, putting them to list_RDR where they are properly formatted
    # adding RDR to database
    def createList(self):
        NewNumberRDR = (self.sheet.cell(row=self.RoomRow, column=6).value)+1
        self.sheet.cell(row=self.RoomRow, column=6).value = NewNumberRDR
        self.book.save('./输入1.xlsx')


        prepareList = self.getValues()
        ThisList = list_RDR(self.RoomID, prepareList)
        OutputScreen = ThisList.ReadyList()
        return OutputScreen


    # Getting values from Database and time method
    def getValues(self):
        OurTimeList = self.getRequestDuration()
        OurTemperature = self.ValuesSource.Temperature
        OurFanSpeed = self.ValuesSource.FanSpeed
        OurFeeRate = self.ValuesSource.NewFeeRate
        OurFee = self.ValuesSource.Fee
        OurDateIn = OurTimeList[0]
        OurDateOut = OurTimeList[1]
        OurDuration = OurTimeList[2]
        result = [str(OurTemperature), str(OurFanSpeed), str(OurFeeRate), str(OurFee), OurDateIn, OurDateOut, OurDuration]
        return result


    # does everything related to time and puts it in special list
    def getRequestDuration(self):
        import datetime
        TimeIn = self.ValuesSource.CheckInTime
        TimeNow = self.ValuesSource.TimeNow
        Duration = TimeNow - TimeIn
        result = [str(TimeIn), str(TimeNow), str(Duration)]
        return result





class list_RDR:              # Gives us the PROPERLY FORMATTED LIST.
    def __init__(self, RoomID, prepareList):
        self.RoomID: int = RoomID
        self.prepareList = prepareList

    def ReadyList(self):
        prepareList = ServiceRDR(self.RoomID)
        prepareList.getValues()
        StringOne = str("Your room:\t\t" + self.RoomID)
        StringTwo = str("Temperature:\t\t" + self.prepareList[0])
        StringThree = str("FanSpeed:\t\t" + self.prepareList[1])
        StringFour = str("Fee Rate:\t\t" + self.prepareList[2])
        StringFive = str("Fee:\t\t" + self.prepareList[3])
        StringSix = ("Check In Time:\t" + self.prepareList[4])
        StringSeven = ("Time Now:\t" + self.prepareList[5])
        StringEight = ("Time in hotel:\t" + self.prepareList[6])
        result = [StringOne, StringTwo, StringThree, StringFour, StringFive, StringSix, StringSeven, StringEight]
        return result     # returns set of PROPERLY FORMATTED STRINGS.





##################################################################################
#                                                                                #
#                             INVOICE RELATED PART                               #
#                                                                                #
##################################################################################


class ServiceInvoice:
    def __init__(self, ID):
        self.RoomID: int = ID
        self.ValuesSource = Database(self.RoomID)

    def createInvoiceData(self):
        import datetime
        TimeIn = str(self.ValuesSource.CheckInTime)
        TimeNow = str(datetime.datetime.now())
        OurFee = str(self.ValuesSource.Fee)
        ThisList = InvoiceData(self.RoomID, TimeIn, TimeNow, OurFee)
        OutputScreen = ThisList.ReadyList()
        return OutputScreen


class InvoiceData:
    def __init__(self, RoomID, TimeIn, TimeNow, OurFee):
        self.RoomID: int = RoomID
        self.TimeIn = TimeIn
        self.TimeNow = TimeNow
        self.OurFee = OurFee

    def ReadyList(self, ):
        StringOne = str("Your room:\t" + self.RoomID)
        StringTwo = ("Check In Time:\t" + self.TimeIn)
        StringThree = ("Time Now:\t" + self.TimeNow)
        StringFour =  ("Fee:\t" + self.OurFee)
        result = (StringOne, StringTwo, StringThree, StringFour)
        return result


##################################################################################
#                                                                                #
#                                    MAIN                                        #
#                                                                                #
##################################################################################
def UiRdrInput():
    RDRorI = '1'
    OutputType = '1'
    RoomID = str(ui.spinID.value())
    ReceptionQuery = ReceptionController(RDRorI, OutputType, RoomID)
    ReceptionQuery.commandCreateList()

def UiRdrPrintInput():
    RDRorI = '1'
    OutputType = '2'
    RoomID = str(ui.spinID.value())
    ReceptionQuery = ReceptionController(RDRorI, OutputType, RoomID)
    ReceptionQuery.commandCreateList()

def UiInvoiceInput():
    RDRorI = ('2')
    OutputType = ('1')
    RoomID = str(ui.spinID.value())
    ReceptionQuery = ReceptionController(RDRorI, OutputType, RoomID)
    ReceptionQuery.commandCreateList()

if __name__ == "__main__":
            # create the app
            app = QtWidgets.QApplication(sys.argv)

            #create the form and init UI
            HotelMenu = QtWidgets.QMainWindow()
            ui = Ui_HotelMenu()
            ui.setupUi(HotelMenu)
            HotelMenu.show()

            #create hook logic
            ui.RDRbutton.clicked.connect(UiRdrInput)
            ui.InvoiceButton.clicked.connect(UiInvoiceInput)
            ui.RDRPrintButton.clicked.connect(UiRdrPrintInput)

            #run main loop
            sys.exit(app.exec_())