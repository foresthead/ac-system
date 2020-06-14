# you can add other cells / values yourself using the same format
import openpyxl

class control:

    def __init__(self):
        self.RoomID = 1  # dummy value
        self.book1 = openpyxl.load_workbook('./ACC_Database.xlsx')  # initializing book1 (ACC_Database)
        self.sheet1 = self.book1.active

        self.book2 = openpyxl.load_workbook('./输入1.xlsx')  # initializing book2 (输入1)
        self.sheet2 = self.book2.active

    def DB_instance(self):
        self.RoomRow = (int(self.RoomID) + 1)  # +1 cuz the first row is the column titles

        # getting values from book1
        self.Temp_highLimit = self.sheet1.cell(row=2, column=6).value
        self.temp_lowLimit = self.sheet1.cell(row=2, column=7).value
        self.Feerate_H = self.sheet1.cell(row=2, column=8).value
        self.Feerate_M = self.sheet1.cell(row=2, column=9).value
        self.Feerate_L = self.sheet1.cell(row=2, column=10).value

        # printing values to see if they are imported right
        print(self.Temp_highLimit, self.temp_lowLimit, self.Feerate_H, self.Feerate_M, self.Feerate_L)

        # dummy values, just to see if we can change the data
        newTempH = 30
        newTempL = 20
        NewFRH = 0.4
        NewFRM = 0.3
        NewFRL = 0.2

        # putting the new values in book1
        self.sheet1.cell(row=2, column=6).value = newTempH
        self.sheet1.cell(row=2, column=7).value = newTempL
        self.sheet1.cell(row=2, column=8).value = NewFRH
        self.sheet1.cell(row=2, column=9).value = NewFRM
        self.sheet1.cell(row=2, column=10).value = NewFRL

        # putting values in book2
        # renewing times of dispatch
        # If you want to put this part in another class, you should also copy book2 initialization from __init__ part
        self.NewDispatch = (self.sheet2.cell(row=self.RoomRow, column=5).value) + 1
        self.sheet2.cell(row=self.RoomRow, column=5).value = self.NewDispatch

        # saving xlsx files to actually see the result
        self.book1.save('./ACC_Database.xlsx')
        self.book2.save('./输入1.xlsx')

if __name__ == "__main__":
    ControllerTest = control()
    ControllerTest.DB_instance()